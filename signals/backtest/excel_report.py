"""Daily-activity Excel report for a BTC (or any) BacktestResult.

Given a BacktestResult and the original OHLCV prices, writes an .xlsx with:

  - **Summary** sheet: aggregate stats (Sharpe, CAGR, MDD, win rate,
    profit factor, trade count, date range, final equity, total return)
  - **Daily Activity** sheet: one row per trading day with:
      date, open, high, low, close, volume
      signal, state, target_position, action
      daily_return_pct, cumulative_return_pct, drawdown_pct
      equity, cash, units_held
      cumulative_buys, cumulative_sells
  - **Trades** sheet: one row per fill (BUY, SELL, COVER, SHORT, STOP)

Requires openpyxl. Conditional formatting highlights buy/sell action
rows so the user can eyeball them fast.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from signals.backtest.engine import BacktestResult


@dataclass
class ExcelReportPaths:
    xlsx_path: Path
    summary_rows: int
    activity_rows: int
    trade_rows: int


def _style_header(ws, n_cols: int) -> None:
    """Bold the header row and freeze top row. openpyxl is imported lazily."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    align = Alignment(horizontal="center", vertical="center")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
    ws.freeze_panes = "A2"


def _highlight_action_rows(ws, action_col_idx: int, n_rows: int) -> None:
    """Tint BUY rows green, SELL rows red."""
    from openpyxl.styles import PatternFill

    buy_fill = PatternFill("solid", fgColor="C6EFCE")
    sell_fill = PatternFill("solid", fgColor="FFC7CE")
    for row in range(2, n_rows + 2):
        action = ws.cell(row=row, column=action_col_idx).value
        if action is None:
            continue
        if "BUY" in str(action):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = buy_fill
        elif "SELL" in str(action) or "COVER" in str(action):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = sell_fill


def build_daily_activity_frame(
    result: BacktestResult,
    prices: pd.DataFrame,
) -> pd.DataFrame:
    """Reconstruct per-day activity from a BacktestResult.

    Joins price data, the engine's signals frame, and the portfolio's
    equity curve into one wide DataFrame indexed by date.
    """
    eq = result.equity_curve.copy()
    if eq.empty:
        return pd.DataFrame()

    # Normalize timestamps — strip tz so merging with OHLC (which may
    # be naive after Excel round-tripping) works consistently.
    def _naive(idx: pd.DatetimeIndex) -> pd.DatetimeIndex:
        return idx.tz_convert("UTC").tz_localize(None) if idx.tz is not None else idx

    activity = pd.DataFrame(index=_naive(pd.DatetimeIndex(eq.index)))
    activity.index.name = "date"
    activity["equity"] = eq.values

    # Attach OHLCV for the overlapping period.
    px = prices.copy()
    px.index = _naive(pd.DatetimeIndex(px.index))
    px = px[~px.index.duplicated(keep="last")]
    for col in ("open", "high", "low", "close", "volume"):
        if col in px.columns:
            activity[col] = px[col].reindex(activity.index)

    # Attach signals, shifted by one bar to align with the action they
    # cause. Engine timing convention is strict no-lookahead:
    #
    #   decision at bar t  →  trade at bar t+1's open  →  equity mark at t+1
    #
    # So on the activity row for bar t+1, the `signal` column must show
    # the decision made at bar t (the decision that caused this row's
    # action), not the new decision made at bar t+1 (which will cause
    # next bar's action). Without this shift, the Excel looks confusing:
    # rows show signal=HOLD alongside action=BUY because the BUY was
    # caused by yesterday's BUY decision while today's decision is HOLD.
    #
    # `next_signal` preserves the "decision made on this bar" info for
    # users who want to see what the model thinks *now* — it's just not
    # the column that explains this row's action.
    sig = result.signals.copy()
    if not sig.empty:
        sig.index = _naive(pd.DatetimeIndex(sig.index))
        # Reindex to activity timeline, then shift forward by one bar.
        # The first activity row ends up with NaN (its action was caused
        # by a warmup-period decision outside the eval window), and the
        # last activity row shows the final pre-flatten decision.
        sig_on_activity = sig.reindex(activity.index)
        sig_aligned = sig_on_activity.shift(1)
        for col in ("signal", "state", "state_label", "confidence",
                    "expected_return", "target_position"):
            if col in sig_aligned.columns:
                activity[col] = sig_aligned[col]
        # Also expose the "decision made on this bar" (no shift) for
        # users who want to see what the model is saying *right now*.
        if "signal" in sig_on_activity.columns:
            activity["next_signal"] = sig_on_activity["signal"]
        if "target_position" in sig_on_activity.columns:
            activity["next_target_position"] = sig_on_activity["target_position"]

    # Reconstruct per-bar position and cash from the trade log.
    # This is independent of the Portfolio's internal state; we replay
    # trades forward so the Excel is self-describing.
    #
    # If the activity window starts AFTER the first trade (common when the
    # caller trimmed equity to an evaluation window but passed the full
    # trade list), pre-apply all pre-window trades so the inherited
    # position/cash state is correct at bar 0 of the activity frame.
    cash = float(result.config.initial_cash)
    units = 0.0
    cum_buys = 0  # cumulative across the FULL trade history (includes warmup)
    cum_sells = 0
    bar_action = []
    bar_cash = []
    bar_units = []
    bar_cum_buys = []
    bar_cum_sells = []

    def _apply_trade(t) -> None:
        nonlocal cash, units, cum_buys, cum_sells
        side = t.side
        fill = t.price
        qty = t.qty
        commission = t.commission
        if side == "BUY":
            cash -= qty * fill + commission
            units += qty
            cum_buys += 1
        elif side == "SELL":
            cash += qty * fill - commission
            units -= qty
            cum_sells += 1
        elif side == "SHORT":
            cash += qty * fill - commission
            units -= qty
        elif side == "COVER":
            cash -= qty * fill + commission
            units += qty

    # Normalize trade timestamps and sort
    def _norm(ts: pd.Timestamp) -> pd.Timestamp:
        ts = pd.Timestamp(ts)
        if ts.tz is not None:
            ts = ts.tz_convert("UTC").tz_localize(None)
        return ts

    sorted_trades = sorted(result.trades, key=lambda t: _norm(t.ts))

    # Pre-apply all trades before the first activity bar (warmup trades)
    first_bar = activity.index[0]
    trade_idx = 0
    while trade_idx < len(sorted_trades) and _norm(sorted_trades[trade_idx].ts).normalize() < first_bar.normalize():
        _apply_trade(sorted_trades[trade_idx])
        trade_idx += 1

    # Capture cumulative counts at start of eval window so we can report
    # cumulative buys/sells WITHIN the eval window in the activity sheet,
    # rather than cumulative-since-start-of-warmup.
    warmup_buys = cum_buys
    warmup_sells = cum_sells

    # Index remaining trades by normalized date for O(n) bar merge
    trades_by_date: dict = {}
    for t in sorted_trades[trade_idx:]:
        key = _norm(t.ts).normalize()
        trades_by_date.setdefault(key, []).append(t)

    for ts in activity.index:
        key = pd.Timestamp(ts).normalize()
        actions_here: list[str] = []
        if key in trades_by_date:
            for t in trades_by_date[key]:
                _apply_trade(t)
                actions_here.append(f"{t.side}({t.qty:.4f}@{t.price:.2f})")
        bar_action.append(" + ".join(actions_here) if actions_here else "")
        bar_cash.append(cash)
        bar_units.append(units)
        # Subtract warmup counts so the sheet shows *eval-window* cumulatives
        bar_cum_buys.append(cum_buys - warmup_buys)
        bar_cum_sells.append(cum_sells - warmup_sells)

    activity["action"] = bar_action
    activity["cash"] = bar_cash
    activity["units_held"] = bar_units
    activity["cumulative_buys"] = bar_cum_buys
    activity["cumulative_sells"] = bar_cum_sells

    # Returns, drawdowns.
    activity["daily_return_pct"] = activity["equity"].pct_change() * 100.0
    first_eq = activity["equity"].iloc[0]
    activity["cumulative_return_pct"] = (activity["equity"] / first_eq - 1.0) * 100.0
    running_max = activity["equity"].cummax()
    activity["drawdown_pct"] = (activity["equity"] / running_max - 1.0) * 100.0

    # Reorder columns for readability. `signal` and `target_position`
    # are the (shifted) decision that CAUSED today's action, so they
    # sit right next to `action` for visual alignment. `next_signal`
    # and `next_target_position` are the decision made ON this bar
    # (driving tomorrow's action); they live at the end so users
    # who care about "what does the model think right now" can still
    # see them without clutter.
    ordered = [
        "open", "high", "low", "close", "volume",
        "state", "confidence", "expected_return",
        "signal", "target_position", "action",
        "units_held", "cash", "equity",
        "daily_return_pct", "cumulative_return_pct", "drawdown_pct",
        "cumulative_buys", "cumulative_sells",
        "next_signal", "next_target_position",
    ]
    present = [c for c in ordered if c in activity.columns]
    activity = activity[present]
    return activity.reset_index()


def build_summary_frame(
    result: BacktestResult,
    symbol: str,
    extra: dict | None = None,
) -> pd.DataFrame:
    m = result.metrics
    bm = result.benchmark_metrics
    eq = result.equity_curve
    first_eq = float(eq.iloc[0]) if len(eq) > 0 else result.config.initial_cash
    last_eq = float(eq.iloc[-1]) if len(eq) > 0 else result.config.initial_cash
    total_return = (last_eq / first_eq - 1.0) * 100.0 if first_eq > 0 else 0.0

    rows = [
        ("Symbol", symbol),
        ("Model", result.config.model_type),
        ("Start date", str(result.start.date())),
        ("End date", str(result.end.date())),
        ("Window length (bars)", len(eq)),
        ("Equity at window start", f"${first_eq:,.2f}"),
        ("Equity at window end", f"${last_eq:,.2f}"),
        ("Total return (window)", f"{total_return:+.2f}%"),
        ("CAGR", f"{m.cagr * 100:+.2f}%"),
        ("Sharpe", f"{m.sharpe:+.3f}"),
        ("Max drawdown", f"{m.max_drawdown * 100:+.2f}%"),
        ("Calmar", f"{m.calmar:+.3f}"),
        ("Win rate", f"{m.win_rate * 100:.1f}%"),
        ("Profit factor", f"{m.profit_factor:.3f}"),
        ("# round-trip trades", m.n_trades),
        ("---", "---"),
        ("Benchmark (B&H) Sharpe", f"{bm.sharpe:+.3f}"),
        ("Benchmark (B&H) CAGR", f"{bm.cagr * 100:+.2f}%"),
        ("Benchmark (B&H) MDD", f"{bm.max_drawdown * 100:+.2f}%"),
        ("Strategy Sharpe - Benchmark Sharpe", f"{m.sharpe - bm.sharpe:+.3f}"),
    ]
    if extra:
        rows.append(("---", "---"))
        for k, v in extra.items():
            rows.append((k, str(v)))
    # Timing convention note — explains the meaning of `signal` vs
    # `next_signal` in the Daily Activity sheet for any user who opens
    # the workbook without the surrounding context.
    rows.append(("---", "---"))
    rows.append((
        "Timing convention",
        "signal = decision that caused today's action (yesterday's decision). "
        "next_signal = decision made today, will cause tomorrow's action.",
    ))
    return pd.DataFrame(rows, columns=["metric", "value"])


def build_trade_frame(
    result: BacktestResult,
    date_range: tuple[pd.Timestamp, pd.Timestamp] | None = None,
) -> pd.DataFrame:
    """Return a trades DataFrame, optionally clipped to a date range.

    If date_range is provided, only trades whose timestamp falls in
    [start, end] are returned. This lets the Excel report surface only
    eval-window trades even when the caller passed the full trade list
    (including warmup trades) for state-reconstruction purposes.
    """
    if not result.trades:
        return pd.DataFrame(
            columns=["ts", "side", "price", "qty", "commission", "pnl", "reason"]
        )
    rows = []
    for t in result.trades:
        ts = pd.Timestamp(t.ts)
        if ts.tz is not None:
            ts = ts.tz_convert("UTC").tz_localize(None)
        if date_range is not None:
            lo, hi = date_range
            if lo is not None and ts < lo:
                continue
            if hi is not None and ts > hi:
                continue
        rows.append({
            "ts": ts,
            "side": t.side,
            "price": t.price,
            "qty": t.qty,
            "commission": t.commission,
            "pnl": t.pnl,
            "reason": t.reason,
        })
    return pd.DataFrame(rows)


def write_excel_report(
    result: BacktestResult,
    prices: pd.DataFrame,
    out_path: str | Path,
    symbol: str = "",
    extra_summary: dict | None = None,
) -> ExcelReportPaths:
    """Write a full daily-activity Excel report. Returns paths and row counts."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    summary = build_summary_frame(result, symbol=symbol or result.symbol, extra=extra_summary)
    activity = build_daily_activity_frame(result, prices)
    # Clip the Trades sheet to the activity window so warmup-period trades
    # (used only for state reconstruction) don't clutter the user's view.
    if len(activity) > 0 and "date" in activity.columns:
        trade_range = (
            pd.Timestamp(activity["date"].iloc[0]),
            pd.Timestamp(activity["date"].iloc[-1]),
        )
    else:
        trade_range = None
    trades = build_trade_frame(result, date_range=trade_range)

    # pandas + openpyxl write
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        activity.to_excel(writer, sheet_name="Daily Activity", index=False)
        trades.to_excel(writer, sheet_name="Trades", index=False)

    # Post-write styling
    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    if "Summary" in wb.sheetnames:
        _style_header(wb["Summary"], n_cols=2)
        wb["Summary"].column_dimensions["A"].width = 36
        wb["Summary"].column_dimensions["B"].width = 28
    if "Daily Activity" in wb.sheetnames:
        ws = wb["Daily Activity"]
        _style_header(ws, n_cols=ws.max_column)
        # Highlight action rows
        cols = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        if "action" in cols:
            _highlight_action_rows(ws, cols["action"], len(activity))
        # Column widths
        for c in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 14
    if "Trades" in wb.sheetnames:
        _style_header(wb["Trades"], n_cols=7)
        for c in range(1, 8):
            wb["Trades"].column_dimensions[wb["Trades"].cell(row=1, column=c).column_letter].width = 14
    wb.save(out_path)

    return ExcelReportPaths(
        xlsx_path=out_path,
        summary_rows=len(summary),
        activity_rows=len(activity),
        trade_rows=len(trades),
    )
